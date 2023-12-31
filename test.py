import openpyxl
import pandas as pd

#置換する文字を返す関数
def text_to_array(cell):
    _str=cell.value
    _str=_str.replace(' ', '')
    arrays=_str.split(',')
    return arrays
# 特定の列を複数検索
def search_columns(column, keyword):
    result = []
    for cell in column:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if keyword in value:
            cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
            result.append(cell_address)
            
    return result



wb = openpyxl.load_workbook("uploads/三者面談（回答）.xlsx")
ws = wb.worksheets[0]

#先生のセルの横軸取得
teacher_cell=ws["C2"]
table_row=text_to_array(ws["C2"])
replace_after=[i for i in range(len(table_row))]
print(replace_after)

print(table_row)
#不要行の削除
ws.delete_cols(3)
ws.delete_cols(1)
ws.delete_rows(2)


#ここで置換を行う

i=1
while True:
    for row in ws.iter_rows(max_row=1 ,min_col=4):
        for cell in row:
            for index,item in enumerate(table_row):
                if cell.offset(i,0).value is None:
                    cell.offset(i,0).value=str(-1)
                    break
                else:
                    cell.offset(i,0).value=cell.offset(i,0).value.replace(item, str(index)).replace(" ","")
    if ws["B2"].offset(i,0).value is None:
        break
    i+=1


# 兄弟の有無を1,0であらわす。
for row in ws.iter_rows(min_col=3,max_col=3,min_row=2):
    for cell in row:
        if cell.value is None:
            break
        cell.value=cell.value.replace('いる', '1').replace('いない', '0')



# #新しいテーブルの横軸作成
new_ws= wb.create_sheet(index=0, title="makedata")
new_ws["A1"].value="出席番号"
new_ws["B1"].value="名前"
new_ws["C1"].value="兄弟"


# 列をループ
for j in range(1,4):
    #行をループ
    for i in range(2,100):
        if ws.cell(row=i, column=j).value is None:
            break
        #値をprint
        new_ws.cell(row=i, column=j).value=ws.cell(row=i, column=j).value

table_column=[]
for row in ws.iter_rows(max_row=1 ,min_col=4):
    for cell in row:
        if cell.value is None:
            break
        table_column.append(cell.value.replace('出席可能日 [', '').replace("]", ""))
print(table_column)

for i in range(1,len(table_column)*len(table_row)+1):
    new_ws["C1"].offset(0,i).value=i

# できた表に1を入力していく。0はPandasで補完
for row in ws.iter_rows(min_row=2 ,min_col=4):
    for cell in row:
        if cell.value is None:
            break
        for i in cell.value.split(","):
            if int(i) == -1:
                break
            new_ws.cell(row=cell.row, column=int(i)+len(replace_after)*(cell.column-len(replace_after))+4).value=1

    if cell.value is None:
        break

df= pd.read_excel('uploads/replace.xlsx', index_col=[0,1,2])
df=df.fillna(0)
df["Total"] = df.sum(axis=1)
df=df.sort_values(by=["兄弟","Total"],ascending=[False,True])
df.loc["Total"] = df.sum()

#0でない行を抽出する
df_process = df.loc[:, df.iloc[0] != 0]
#その中で最小のIndexを取得する
print(df_process.iloc[-1].idxmin())

put_num=df_process.iloc[-1].idxmin()

#テーブルの作成
print_ws= wb.create_sheet(index=2, title="print")
for i,value in enumerate(table_row):
    print_ws.cell(row=2+i,column=1).value=value
for i,value in enumerate(table_column):
    print_ws.cell(row=1,column=2+i).value=value

#書き込む行と列を指定
q, mod = divmod(put_num, len(table_column))
print(q)
print(mod)
print_ws.cell(row=2+q,column=1+mod).value=1

df=df.drop(df.index[[0]])
print(df)

ws.title="replace"



wb.save('uploads/replace.xlsx')
wb.close()