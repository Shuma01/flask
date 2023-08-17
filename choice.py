import pandas as pd
import data_cleansing as dc
import openpyxl

#dfファイルの作成
def make_df_file():
    df=pd.read_excel('uploads/replace.xlsx', index_col=[0,1,2])
    df=df.fillna(0)
    df.to_excel('uploads/replace.xlsx', sheet_name='makedata')

#最初にdfの作成
def make_df():
    df=pd.read_excel('uploads/replace.xlsx', index_col=[0,1,2])
    return df


#合計数の少ない人でほかの人のかぶりが少ない日の数を返す
def total_min(df):
    df.loc["Total"] = df.sum()
    df = df.loc[:, df.iloc[0] != 0]
    min_day=df.iloc[-1].idxmin()
    df=df.drop(df.index[[-1]])
    return min_day

#集計
def add_total_df(df):
    df["Total"] = df.sum(axis=1)
    df=df.sort_values(by=["兄弟","Total"],ascending=[False,True])
    df=df.drop(df.columns[[-1]], axis=1)
    return df

#出席可能な日を求める(番号)
def enable_day(df):
    df = df.loc[:, df.iloc[0] != 0]
    return df.columns.to_list()[:-1]

#番号から日を求める
def num_to_day(day_num,table_row,table_column):
    q, mod = divmod(day_num, len(table_row))
    if mod is None:
        mod=0
    return[table_column[q],table_row[mod-1]]

#日から番号を求める
def day_to_num(day_data,table_row,table_column):
    q=table_column.index(day_data[0])
    mod=table_row.index(day_data[1])
    return q*len(table_row)+mod+1

def make_excel(table_row,table_column):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i,value in enumerate(table_row):
        ws.cell(row=2+i,column=1).value=value
    for i,value in enumerate(table_column):
        ws.cell(row=1,column=2+i).value=value
    
    wb.save('uploads/print.xlsx')
    wb.close()

def write_data(day_data,df,table_row,table_column):
    wb = openpyxl.load_workbook("uploads/print.xlsx")
    ws= wb.worksheets[0]
    column=table_column.index(day_data[0])+2
    row=table_row.index(day_data[1])+2
    ws.cell(row=row,column=column).value=df.index[0][1]
    wb.save("uploads/print.xlsx")
    wb.close()


def delete_data(day_data,df,table_row,table_column):
    del_num=day_to_num(day_data,table_row,table_column)
    df=df.drop(del_num,axis=1)
    df=df.drop(df.index[[0]])
    df.to_excel('uploads/replace.xlsx', sheet_name='makedata')
    return df

#兄弟がいるか調べる
def is_brother(df):
    if int(df.index[0][2])==1:
        return True
    else:
        return False
    

def print_table(table_row,table_column):
    wb = openpyxl.load_workbook("uploads/print.xlsx")
    ws= wb.worksheets[0]
    rows=[]
    for row in ws.iter_rows(min_row=2 ,max_row=len(table_row)+1,min_col=1,max_col=len(table_column)+1):
        rows.append(row)
    return rows

# table_column=['１日目', '２日目', '３日目']
# table_row=['9:00-10:00', '10:00-11:00', '11:00-12:00', '12:00-13:00']

