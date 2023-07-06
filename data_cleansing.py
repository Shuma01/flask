import openpyxl

#置換する文字を返す関数
def text_to_array(cell):
    _str=cell.value
    _str=_str.replace(' ', '')
    arrays=_str.split(',')
    return arrays


wb = openpyxl.load_workbook("uploads/三者面談（回答）.xlsx")
ws = wb.worksheets[0]
ws.title="replace"

#先生のセルの横軸取得
teacher_cell=ws["C2"]
table_row=text_to_array(ws["C2"])
replace_after=[i for i in range(len(table_row))]
print(replace_after)

print(table_row)
#不要行の削除
ws.delete_cols(3)
ws.delete_cols(1)
ws.delete_rows(2)


#ここで置換を行う

i=1
while True:
    for row in ws.iter_rows(max_row=1 ,min_col=4):
        for cell in row:
            for index,item in enumerate(table_row):
                if cell.offset(i,0).value is None:
                    cell.offset(i,0).value=str(-1)
                    break
                else:
                    cell.offset(i,0).value=cell.offset(i,0).value.replace(item, str(index)).replace(" ","")
    if ws["B2"].offset(i,0).value is None:
        break
    i+=1


# 兄弟の有無を1,0であらわす。
for row in ws.iter_rows(min_col=3,max_col=3,min_row=2):
    for cell in row:
        if cell.value is None:
            break
        cell.value=cell.value.replace('いる', '1').replace('いない', '0')



# #新しいテーブルの横軸作成
new_ws= wb.create_sheet(index=0, title="makedata")
new_ws["A1"].value="出席番号"
new_ws["B1"].value="名前"
new_ws["C1"].value="兄弟"


# 列をループ
for j in range(1,4):
    #行をループ
    for i in range(2,100):
        if ws.cell(row=i, column=j).value is None:
            break
        #値をprint
        new_ws.cell(row=i, column=j).value=ws.cell(row=i, column=j).value

table_column=[]
for row in ws.iter_rows(max_row=1 ,min_col=4):
    for cell in row:
        if cell.value is None:
            break
        table_column.append(cell.value.replace('出席可能日 [', '').replace("]", ""))
print(table_column)

for i in range(1,len(table_column)*len(table_row)+1):
    new_ws["C1"].offset(0,i).value=i

# できた表に1を入力していく。0はPandasで補完
for row in ws.iter_rows(min_row=2 ,min_col=4):
    for cell in row:
        if cell.value is None:
            break
        for i in cell.value.split(","):
            if int(i) == -1:
                break
            new_ws.cell(row=cell.row, column=int(i)+len(replace_after)*(cell.column-len(replace_after))+4).value=1

    if cell.value is None:
        break
wb.save('uploads/replace.xlsx')
wb.close()