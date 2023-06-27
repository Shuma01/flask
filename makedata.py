import openpyxl
# 特定の列を検索
def search_column(column, keyword):
    for cell in column:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if value in keyword:
            result= openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
    return result

# 特定の列を複数検索
def search_columns(column, keyword):
    result = []
    for cell in column:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if keyword in value:
            cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
            result.append(cell_address)
            
    return result

# 特定の行を検索
def search_row(row, keyword):
    for cell in row:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if value == keyword:
            result = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)

    return result



wb = openpyxl.load_workbook("replace.xlsx")
ws = wb.worksheets[0]
new_ws= wb.create_sheet(index=0, title="makedata")
student_number=search_column(ws['1'],"出席番号")
student_name=search_column(ws['1'],"名前を入力")
student_brother=search_column(ws['1'],"兄弟はいますか")
new_ws["A1"].value=ws[student_number].value
new_ws["B1"].value=ws[student_name].value
new_ws["C1"].value=ws[student_brother].value


#先生と書かれたセルの取得
teacher_cell=search_row(ws['B'], "先生")


#置換する文字を返す関数
def generateReplace(cell):
    _time_str=ws[cell].offset(0,1).value
    _time_str=_time_str.replace(' ', '')
    time_arrays=_time_str.split(',')
    return time_arrays


replace_items=generateReplace(teacher_cell)

move_row=0
day_rows=search_columns(ws["1"], "出席可能")
for cell in day_rows:
    for replace_item in range(len(replace_items)):
        time_table=ws[cell].value.replace('出席可能日 [', '').replace("]", "")
        ws[cell].value=ws[cell].value.replace('出席可能日 [', '').replace("]", "")
        time_table+=","+str(replace_item)
        new_ws["D1"].offset(0,move_row).value=time_table
        move_row+=1


ws.delete_cols(3)
ws.delete_cols(1)
ws.delete_rows(2)
wb.save("replace.xlsx")